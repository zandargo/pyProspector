"""
PyProspector — B2B Lead Prospecting Tool via Google Maps
=========================================================
Stack : Python · Playwright · playwright-stealth · Streamlit · Pandas · openpyxl

Scraping flow (2 phases):
  Phase 1 → Searches "{niche} {city}" on Maps, scrolls the feed and collects place URLs
  Phase 2 → Visits each URL individually and extracts the establishment details

Fitness score for web dev / digital marketing services:
  base  = (rating × 10) × (number_of_reviews / 100)
  final = base × 10   →  no website        (🔥 highest priority)
  final = base × 5    →  social network URL (🟡 medium priority)
  final = base × 1    →  proper website     (low priority)
"""

from __future__ import annotations

import asyncio
import io
import random
import re
import sys
import time
from urllib.parse import quote_plus, urljoin



import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import TimeoutError as PWTimeout
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

# ── Page configuration ───────────────────────────────────────────────────────
st.set_page_config(
    page_title="PyProspector – B2B Leads",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Constants ────────────────────────────────────────────────────────────────
MAPS_SEARCH_URL = "https://www.google.com/maps/search/"
MAPS_BASE_URL   = "https://www.google.com"

# Pool of real user-agents for rotation and reduced detection
USER_AGENTS = [
    (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/132.0.0.0 Safari/537.36"
    ),
    (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 Safari/537.36"
    ),
    (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) "
        "Gecko/20100101 Firefox/133.0"
    ),
]


# ══════════════════════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def _delay(min_s: float = 0.8, max_s: float = 2.2) -> None:
    """Random pause to simulate human rhythm and avoid rate limiting."""
    time.sleep(random.uniform(min_s, max_s))


def _parse_float(text: str) -> float:
    """Safely converts '4,6' or '4.6' to float."""
    cleaned = text.strip().replace(",", ".")
    m = re.search(r"\d+\.?\d*", cleaned)
    return float(m.group()) if m else 0.0


def _parse_int(text: str) -> int:
    """Extracts the first sequence of digits from a string."""
    digits = re.sub(r"[^\d]", "", text)
    return int(digits) if digits else 0


def _safe_text(locator) -> str:
    """Returns inner_text of a Playwright locator or '' on error."""
    try:
        return locator.inner_text(timeout=3_000).strip()
    except Exception:
        return ""


def _safe_attr(locator, attr: str) -> str:
    """Returns an attribute of a Playwright locator or '' on error."""
    try:
        return (locator.get_attribute(attr, timeout=3_000) or "").strip()
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════════════
# SCRAPING — helper functions
# ══════════════════════════════════════════════════════════════════════════════

def _handle_consent(page) -> None:
    """Accepts Google cookie consent banners if present."""
    for text in ["Aceitar tudo", "Accept all", "Concordo", "Agree"]:
        try:
            btn = page.get_by_role("button", name=re.compile(text, re.I)).first
            if btn.is_visible(timeout=2_000):
                btn.click()
                _delay(1.0, 2.0)
                return
        except Exception:
            pass


def _collect_place_urls(page, target: int) -> list[str]:
    """
    Scrolls the Maps results feed and collects unique place URLs.

    Strategy:
    - Finds anchors whose href contains '/maps/place/'
    - Strips tracking query parameters (keeps only the base path)
    - Stops scrolling when `target` URLs are reached or after 8 iterations with no new results

    Returns a list of absolute URLs (at most `target` items).
    """
    seen: set[str] = set()
    no_new_iters = 0

    while len(seen) < target and no_new_iters < 8:
        anchors = page.query_selector_all(
            "div[role='feed'] a[href*='/maps/place/']"
        )
        prev_len = len(seen)

        for anchor in anchors:
            href = anchor.get_attribute("href") or ""
            if not href:
                continue
            # Ensure the URL is absolute
            full_url = (
                href if href.startswith("http")
                else urljoin(MAPS_BASE_URL, href)
            )
            # Strip query string parameters (tracking, etc.)
            clean_url = re.split(r"\?", full_url)[0]
            seen.add(clean_url)

        if len(seen) == prev_len:
            no_new_iters += 1
        else:
            no_new_iters = 0

        # Scroll the feed to load more results
        try:
            page.evaluate(
                "document.querySelector(\"div[role='feed']\").scrollBy(0, 1200)"
            )
        except Exception:
            break

        _delay(1.5, 2.8)

    return list(seen)[:target]


def _extract_place_data(page, maps_url: str = "") -> dict | None:
    """
    Extracts all relevant fields from a Google Maps place page.

    Uses a single JavaScript scan of all [data-item-id] elements for
    contact info (more reliable than individual Playwright queries).

    Returns a dict with the fields or None if the name cannot be extracted.
    """
    # Wait for the main title to load
    try:
        page.wait_for_selector("h1", timeout=10_000)
    except PWTimeout:
        return None

    # ── Name ──────────────────────────────────────────────────────────────────
    name = _safe_text(page.locator("h1").first)
    if not name:
        return None

    # ── Category ──────────────────────────────────────────────────────────────
    category = ""
    for sel in [
        "button.DkEaL",
        "span.YkuOqf",
        "div.LBgpqf button",
        "[class*='fontBodyMedium'] button",
    ]:
        try:
            el = page.query_selector(sel)
            if el:
                category = el.inner_text().strip()
                break
        except Exception:
            pass

    # ── Rating and Reviews ─────────────────────────────────────────────────────
    # The 'F7nice' block typically contains: span(4.6) + span((1,234))
    rating  = 0.0
    reviews = 0
    try:
        nice_block = page.query_selector("div.F7nice")
        if nice_block:
            for sp in nice_block.query_selector_all("span"):
                txt = sp.inner_text().strip()
                if re.match(r"^\d[,\.]\d$", txt):
                    rating = _parse_float(txt)
                elif re.search(r"\([\d\s\.,]+\)", txt):
                    reviews = _parse_int(re.sub(r"[^\d]", "", txt))
    except Exception:
        pass

    # ── Contact info via JavaScript scan ──────────────────────────────────────
    # Wait for the info panel buttons to appear, then read everything in one JS
    # call — avoids races and glyph-character issues from inner_text().
    try:
        page.wait_for_selector("[data-item-id]", timeout=8_000)
    except PWTimeout:
        pass

    info_items: dict = {}
    try:
        info_items = page.evaluate(
            """
            () => {
                const result = {};
                document.querySelectorAll('[data-item-id]').forEach(el => {
                    const id    = el.getAttribute('data-item-id') || '';
                    const label = (el.getAttribute('aria-label') || '').trim();
                    const href  = el.getAttribute('href') || '';
                    // innerText of the text-only child div (skips SVG / icon spans)
                    const textDiv = el.querySelector(
                        'div.Io6YTe, div[class*="fontBody"], div.rogA2c + div'
                    );
                    const text = textDiv
                        ? textDiv.innerText.split('\\n')[0].trim()
                        : el.innerText.split('\\n')[0].trim();
                    if (id) result[id] = { label, text, href };
                });
                return result;
            }
            """
        )
    except Exception:
        pass

    def _best(d: dict) -> str:
        """Return the best non-empty string from label → text → href."""
        return (d.get("label") or d.get("text") or d.get("href") or "").strip()

    def _strip_glyphs(s: str) -> str:
        return re.sub(r"[\x00-\x1f\x7f\ue000-\uf8ff]+", "", s).strip()

    # ── Address ───────────────────────────────────────────────────────────────
    address = ""
    addr_raw = _strip_glyphs(_best(info_items.get("address", {})))
    if addr_raw:
        addr_raw = re.sub(r"^[Ee]ndere[çc]o[:\s]*", "", addr_raw)
        addr_raw = re.sub(r"^[Aa]ddress[:\s]*", "", addr_raw).strip()
        address = addr_raw

    # ── Phone ─────────────────────────────────────────────────────────────────
    phone = ""
    for key, val in info_items.items():
        if key.startswith("phone:tel:"):
            raw = _strip_glyphs(_best(val))
            raw = re.sub(r"^[Tt]elefone[:\s]*", "", raw)
            raw = re.sub(r"^[Pp]hone[\s\S]*?:\s*", "", raw)
            raw = re.sub(r"^[Cc]all[\s\S]*?:\s*", "", raw)
            raw = raw.strip()
            if raw:
                phone = raw
            break

    # ── Website ───────────────────────────────────────────────────────────────
    website = ""
    auth = info_items.get("authority", {})
    website = (
        auth.get("href")
        or _strip_glyphs(auth.get("label") or auth.get("text") or "")
    )

    return {
        "name":     name,
        "category": category,
        "address":  address,
        "phone":    phone,
        "website":  website,
        "rating":   rating,
        "reviews":  reviews,
        "maps_url": maps_url,
    }


def _extract_reviews(page, max_reviews: int) -> list[str]:
    """
    Switches to the Reviews tab on a Maps place page and returns the plain
    text of the first `max_reviews` customer reviews.

    Returns an empty list if no reviews are found or on any error.
    """
    # Try to click the Reviews tab to make all reviews visible
    for sel in [
        "button[aria-label*='review' i]",
        "button[data-tab-index='1']",
    ]:
        try:
            el = page.query_selector(sel)
            if el and el.is_visible(timeout=2_000):
                el.click()
                _delay(1.0, 2.0)
                break
        except Exception:
            pass

    # Wait for review text elements to appear
    try:
        page.wait_for_selector("span.wiI7pd", timeout=8_000)
    except PWTimeout:
        return []

    # Expand truncated reviews ("More" buttons) for the first N items
    try:
        more_btns = page.query_selector_all("button.w8nwRe, button[aria-label*='more' i]")
        for btn in more_btns[:max_reviews]:
            try:
                if btn.is_visible(timeout=1_000):
                    btn.click()
                    _delay(0.15, 0.35)
            except Exception:
                pass
    except Exception:
        pass

    # Collect review texts
    texts: list[str] = []
    try:
        spans = page.query_selector_all("span.wiI7pd")
        for sp in spans[:max_reviews]:
            try:
                t = sp.inner_text().strip()
                if t:
                    texts.append(t)
            except Exception:
                pass
    except Exception:
        pass

    return texts


# ══════════════════════════════════════════════════════════════════════════════
# SCRAPING — main function
# ══════════════════════════════════════════════════════════════════════════════

def scrape_google_maps(
    niche: str,
    city: str,
    max_results: int = 50,
    min_rating: float = 0.0,
    review_keywords: list[str] | None = None,
    review_scan_count: int = 5,
    progress_callback=None,
    status_callback=None,
) -> list[dict]:
    """
    Scrapes establishments from Google Maps using Playwright + stealth.

    Two-phase flow:
      1. Opens the search, scrolls the feed and collects place URLs.
      2. Visits each URL individually and extracts name, category, address,
         phone, website, rating and number of reviews.

    Args:
        niche:              Business segment (e.g. "dentists").
        city:               Location (e.g. "São Paulo, Brasil").
        max_results:        Maximum number of leads to extract.
        min_rating:         Minimum rating filter (0.0 = no filter).
        review_keywords:    Optional list of keywords to search in reviews.
        review_scan_count:  How many reviews to read per place (1‑200).
        progress_callback:  Callable(current: int, total: int) for progress.

    Returns:
        List of dicts, one per lead found.
    """
    _kw_list: list[str] = [kw.strip() for kw in (review_keywords or []) if kw.strip()]
    leads: list[dict] = []
    query      = f"{niche} {city}"
    search_url = MAPS_SEARCH_URL + quote_plus(query)
    ua         = random.choice(USER_AGENTS)

    def _status(msg: str) -> None:
        if status_callback:
            status_callback(msg)

    # Streamlit sets WindowsSelectorEventLoopPolicy on Windows, which breaks
    # Playwright's subprocess launch. Force ProactorEventLoop right before
    # sync_playwright() creates its internal event loop.
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

    _status("🌐 Launching browser…")

    with sync_playwright() as pw:
        # Command-line arguments to reduce automation fingerprints
        browser = pw.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
                "--disable-infobars",
                "--disable-extensions",
                "--window-size=1366,768",
            ],
        )

        ctx = browser.new_context(
            user_agent=ua,
            viewport={"width": 1366, "height": 768},
            locale="en-US",
            timezone_id="America/New_York",
        )
        page = ctx.new_page()

        # Apply stealth patches (removes navigator.webdriver, etc.)
        Stealth().apply_stealth_sync(page)

        # Block images and fonts to reduce bandwidth usage and latency
        page.route(
            re.compile(r"\.(png|jpg|jpeg|gif|webp|svg|woff2?|ttf|otf)(\?|$)", re.I),
            lambda route, _req: route.abort(),
        )

        try:
            # ── PHASE 1: Collect place URLs ────────────────────────────────────
            _status(f"🔍 Phase 1 — Searching Google Maps for **{query}**…")
            page.goto(search_url, wait_until="domcontentloaded", timeout=30_000)
            _delay(2.5, 4.0)
            _handle_consent(page)

            try:
                page.wait_for_selector("div[role='feed']", timeout=20_000)
            except PWTimeout:
                st.warning(
                    "The results panel did not load. "
                    "Check the niche/city or try again in a moment."
                )
                return leads

            _status("📜 Phase 1 — Scrolling feed to collect place links…")
            place_urls = _collect_place_urls(page, max_results)
            if not place_urls:
                st.warning("No results found for this search.")
                return leads

            _status(
                f"🔗 Phase 1 complete — {len(place_urls)} places found. "
                "Starting detail extraction…"
            )
            # ── PHASE 2: Visit each place and extract details ─────────────────
            for i, url in enumerate(place_urls):
                if i > 0:
                    _delay(1.2, 2.8)   # be respectful between requests

                # Extract a human-readable name from the Maps URL path
                _url_name_match = re.search(r"/maps/place/([^/@]+)", url)
                _url_name = (
                    _url_name_match.group(1).replace("+", " ").replace("%20", " ")
                    if _url_name_match else url
                )
                _status(
                    f"🧩 Phase 2 — {i + 1} / {len(place_urls)} — "
                    f"Visiting **{_url_name}**…"
                )

                try:
                    page.goto(url, wait_until="domcontentloaded", timeout=25_000)
                    _delay(1.5, 3.0)

                    data = _extract_place_data(page, maps_url=url)
                    if data is None:
                        continue

                    # Apply minimum rating filter if configured
                    if min_rating > 0.0 and data["rating"] < min_rating:
                        continue

                    # ── Review keyword search (optional) ──────────────────
                    if _kw_list:
                        _status(
                            f"🧩 Phase 2 — {i + 1} / {len(place_urls)} — "
                            f"Reading reviews: **{data['name']}**…"
                        )
                        review_texts = _extract_reviews(page, review_scan_count)
                        found = [
                            kw for kw in _kw_list
                            if any(kw.lower() in rt.lower() for rt in review_texts)
                        ]
                        data["matched_keywords"] = ", ".join(found)
                    else:
                        data["matched_keywords"] = ""

                    leads.append(data)
                    _status(
                        f"🧩 Phase 2 — {i + 1} / {len(place_urls)} — "
                        f"Processing: **{data['name']}**"
                    )

                    if progress_callback:
                        progress_callback(len(leads), len(place_urls))

                except PWTimeout:
                    continue   # page took too long → skip
                except Exception:
                    continue   # any other error → skip

        except Exception as exc:
            st.error(f"Unexpected error during scraping: {exc}")

        finally:
            ctx.close()
            browser.close()

    return leads


# ══════════════════════════════════════════════════════════════════════════════
# DATA PROCESSING
# ══════════════════════════════════════════════════════════════════════════════

def process_data(raw: list[dict]) -> pd.DataFrame:
    """
    Converts the raw leads list into an enriched DataFrame.

    Operations:
    - Type normalisation (rating → float, reviews → int)
    - site_type column ("Page" | "Social Network" | "")
    - Fitness score calculation for web dev / digital marketing
    - Descending sort by score

    Score:
        base  = (rating × 10) × (reviews / 100)
        final = base × 10  →  no website        (🔥 highest priority)
        final = base × 5   →  social network URL (🟡 medium priority)
        final = base × 1   →  proper website     (low priority)
    """
    if not raw:
        return pd.DataFrame(
            columns=[
                "name", "category", "address", "phone", "whatsapp_url",
                "website", "site_type", "matched_keywords",
                "rating", "reviews", "score", "maps_url",
            ]
        )

    df = pd.DataFrame(raw)

    # Normalise types
    df["rating"]  = pd.to_numeric(df["rating"],  errors="coerce").fillna(0.0)
    df["reviews"] = (
        pd.to_numeric(df["reviews"], errors="coerce").fillna(0).astype(int)
    )
    df["website"]          = df.get("website", "").fillna("").str.strip()
    df["matched_keywords"] = df.get("matched_keywords", "").fillna("")

    # Classify site type: "Page", "Social Network", or "" (none)
    _SOCIAL_RE = re.compile(
        r"(instagram|facebook|linkedin)\.com", re.I
    )

    def _classify_site(url: str) -> str:
        if not url:
            return ""
        if _SOCIAL_RE.search(url):
            return "Social Network"
        return "Page"

    df["site_type"] = df["website"].apply(_classify_site)

    # WhatsApp Web link — put formatted phone in fragment so it can be shown
    # as display text. Browsers strip the fragment before contacting wa.me,
    # so the link still opens on the correct number.
    def _whatsapp_url(p: str) -> str:
        phone_str = str(p).strip()
        digits = re.sub(r"[^\d]", "", phone_str)
        if not digits:
            return ""
        return f"https://wa.me/{digits}#{phone_str}" if phone_str else f"https://wa.me/{digits}"

    df["whatsapp_url"] = df["phone"].apply(_whatsapp_url)

    # Score calculation: x10 no site, x5 social network, x1 real page
    base = (df["rating"] * 10) * (df["reviews"] / 100)
    _multiplier = df["site_type"].map({"": 10.0, "Social Network": 5.0, "Page": 1.0})
    df["score"] = (base * _multiplier).round(2)

    # Sort by score (highest priority first)
    df = df.sort_values("score", ascending=False).reset_index(drop=True)

    # Column display order
    ordered = [
        "name", "category", "address", "phone", "whatsapp_url",
        "website", "site_type", "matched_keywords",
        "rating", "reviews", "score", "maps_url",
    ]
    return df[[c for c in ordered if c in df.columns]]


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════════════════════

# Column name mapping for Excel / TSV headers
_COL_LABELS = {
    "name":          "Name",
    "category":      "Category",
    "address":       "Address",
    "phone":         "Phone",
    "whatsapp_url":  "WhatsApp",
    "website":       "Website",
    "site_type":          "Site type",
    "matched_keywords":   "Keywords Found",
    "rating":        "Rating (★)",
    "reviews":       "No. of Reviews",
    "score":         "Score",
    "maps_url":      "Google Maps",
}


def generate_excel(df: pd.DataFrame) -> bytes:
    """
    Generates a formatted Excel (.xlsx) file with openpyxl.

    Formatting:
    - Header: blue background (#2E75B6) + bold + white text
    - Rows WITHOUT a website: light green background (#C6EFCE) — signals opportunity
    - Header frozen at row 1
    - Column widths auto-adjusted to content
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "B2B Leads"

    header_fill    = PatternFill("solid", fgColor="2E75B6")
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    no_site_fill   = PatternFill("solid", fgColor="C6EFCE")   # no website  → green
    social_fill    = PatternFill("solid", fgColor="FFEB9C")   # social only → yellow
    c_align        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_align        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    headers = [_COL_LABELS.get(c, c) for c in df.columns]

    # Header row
    for col_i, header in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col_i, value=header)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = c_align

    # Data rows
    for row_i, row in enumerate(df.itertuples(index=False), 2):
        site_type = getattr(row, "site_type", "Page")
        if site_type == "":
            row_fill = no_site_fill
        elif site_type == "Social Network":
            row_fill = social_fill
        else:
            row_fill = None

        for col_i, val in enumerate(row, 1):
            cell           = ws.cell(row=row_i, column=col_i, value=val)
            cell.alignment = l_align
            if row_fill:
                cell.fill = row_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_w = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 60)

    ws.freeze_panes = "A2"   # freeze header when scrolling

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_txt(df: pd.DataFrame) -> bytes:
    """
    Generates a plain-text list of leads, one labeled field per line.

    Places are separated by two blank lines.
    """
    entries: list[str] = []
    for _, row in df.iterrows():
        lines = []
        for col, label in _COL_LABELS.items():
            if col not in df.columns:
                continue
            val = row[col]
            if col == "site_type":
                val = val if val else "None"
            lines.append(f"{label}: {val}")
        entries.append("\n".join(lines))
    return "\n\n\n".join(entries).encode("utf-8")


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT INTERFACE
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    # ── Session state ─────────────────────────────────────────────────────────
    if "results_df" not in st.session_state:
        st.session_state["results_df"] = None
    if "results_slug" not in st.session_state:
        st.session_state["results_slug"] = ""

    # ── Header ────────────────────────────────────────────────────────────────
    c_logo, c_title = st.columns([1, 9])
    with c_logo:
        st.markdown("## 🎯")
    with c_title:
        st.title("PyProspector")
        st.caption(
            "B2B Lead Prospecting via Google Maps · "
            "Playwright + Streamlit · Ethical Scraping"
        )
    st.divider()

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.header("⚙️ Search Parameters")
        st.markdown("---")

        niche = st.text_input(
            "🏷️ Niche / Segment",
            placeholder="e.g. dentists, clinics, lawyers",
            help="Type of business you want to prospect.",
        )
        city = st.text_input(
            "📍 City / Country",
            placeholder="e.g. New York, USA",
            help="Location for the Google Maps search.",
        )
        max_results = st.slider(
            "🔢 Max. results",
            min_value=5,
            max_value=100,
            value=50,
            step=5,
        )
        min_rating = st.slider(
            "⭐ Minimum rating  (0 = all)",
            min_value=0.0,
            max_value=5.0,
            value=0.0,
            step=0.5,
        )

        st.markdown("---")
        st.markdown("#### 🔎 Review Keyword Search *(optional)*")
        kw_raw = st.text_area(
            "Keywords (one per line)",
            placeholder="e.g.\ndelivery\nbad service\nparking",
            help=(
                "Scan the first N customer reviews of each place. "
                "Matched keywords will appear in the results table."
            ),
            height=100,
        )
        review_scan_count = st.slider(
            "Reviews to scan per place",
            min_value=1,
            max_value=200,
            value=5,
            step=1,
            help="Higher values are more thorough but slower.",
            disabled=not kw_raw.strip(),
        )
        review_keywords = [
            ln.strip() for ln in kw_raw.splitlines() if ln.strip()
        ]

        st.markdown("---")
        run = st.button(
            "🚀 Prospect Leads",
            type="primary",
            width='stretch',
            disabled=not (niche.strip() and city.strip()),
        )
        if not (niche.strip() and city.strip()):
            st.caption("⬆ Fill in the niche and city to enable.")

    # ── Results area ──────────────────────────────────────────────────────────
    if run:
        niche = niche.strip()
        city  = city.strip()

        # Clear previous results before a new search
        st.session_state["results_df"]   = None
        st.session_state["results_slug"] = ""

        info_box = st.info(
            f"🔍 Searching for **{niche}** in **{city}** "
            f"— up to **{max_results}** results…"
        )
        progress = st.progress(0, text="Starting scraper…")
        status   = st.empty()

        def on_status(msg: str) -> None:
            status.markdown(msg)

        def on_progress(current: int, total: int) -> None:
            pct = min(current / max(total, 1), 1.0)
            progress.progress(pct, text=f"Phase 2 — {current}/{total} leads extracted…")

        # Run the scraper (may take a few minutes)
        try:
            collected = scrape_google_maps(
                niche=niche,
                city=city,
                max_results=max_results,
                min_rating=min_rating,
                review_keywords=review_keywords,
                review_scan_count=review_scan_count,
                progress_callback=on_progress,
                status_callback=on_status,
            )
        except Exception as exc:
            st.error(f"Critical error: {exc}")
            st.stop()
        finally:
            progress.empty()
            status.empty()
            info_box.empty()

        if not collected:
            st.warning(
                "No leads were collected. Suggestions:\n"
                "- Check the spelling of the niche and city\n"
                "- Lower the minimum rating\n"
                "- Google Maps may have temporarily throttled requests — try again"
            )
            st.stop()

        df   = process_data(collected)
        slug = (
            f"leads_{niche.replace(' ', '_')}_"
            f"{city.split(',')[0].strip().replace(' ', '_')}"
        )
        st.session_state["results_df"]   = df
        st.session_state["results_slug"] = slug

    # ── Display results (from this run or preserved in session state) ──────────
    if st.session_state["results_df"] is not None:
        df   = st.session_state["results_df"]
        slug = st.session_state["results_slug"]

        # ── Quick metrics ─────────────────────────────────────────────────────
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("📋 Leads collected",  len(df))
        m2.metric("🚫 No website",       int((df["site_type"] == "").sum()))
        m3.metric("⭐ Avg. rating",       f"{df['rating'].mean():.1f}")
        m4.metric("🏆 Top score",        f"{df['score'].max():.1f}")

        st.success(f"✅ {len(df)} leads processed and sorted by score!")
        st.divider()

        # ── Interactive table ──────────────────────────────────────────────────
        st.subheader("📊 Results (sorted by score)")

        hide_pages = st.checkbox("🚫 Hide businesses with a normal webpage", value=False)
        display_df = df[df["site_type"] != "Page"] if hide_pages else df

        max_score = display_df["score"].max() or 1.0
        col_cfg = {
            "name":        st.column_config.TextColumn("Name",        width="medium"),
            "category":    st.column_config.TextColumn("Category",    width="small"),
            "address":     st.column_config.TextColumn("Address",     width="large"),
            "whatsapp_url": st.column_config.LinkColumn(
                               "Phone", width="small",
                               display_text=r"#(.+)",
                               help="Click to open WhatsApp Web for this number",
                           ),
            "website":     st.column_config.LinkColumn("Website",     width="medium"),
            "site_type":        st.column_config.TextColumn("Site type",      width="small",
                                    help="Page · Social Network · (empty = no website)"),
            "matched_keywords": st.column_config.TextColumn("Keywords Found", width="medium",
                                    help="Review keywords matched for this place"),

            "rating":      st.column_config.NumberColumn("★ Rating",  format="%.1f"),
            "reviews":     st.column_config.NumberColumn("Reviews"),
            "score":       st.column_config.ProgressColumn(
                               "Score", max_value=max_score, format="%.1f"
                           ),
            "maps_url":    st.column_config.LinkColumn("📍 Maps",     width="small"),
        }

        # column_order controls which columns are visible; 'phone' (raw text)
        # is kept in the DataFrame for exports but hidden from the table.
        _display_order = [
            "name", "category", "address", "whatsapp_url",
            "website", "site_type", "matched_keywords",
            "rating", "reviews", "score", "maps_url",
        ]

        st.dataframe(
            display_df,
            width='stretch',
            hide_index=True,
            column_config=col_cfg,
            column_order=_display_order,
            height=520,
        )

        # ── Download buttons ───────────────────────────────────────────────────
        st.divider()
        st.subheader("⬇️ Export data")

        dl1, dl2 = st.columns(2)

        with dl1:
            st.download_button(
                label="📥 Download Excel (.xlsx)",
                data=generate_excel(df.drop(columns=["whatsapp_url"], errors="ignore")),
                file_name=f"{slug}.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
                width='stretch',
            )

        with dl2:
            st.download_button(
                label="📄 Download Plain Text Table (.txt)",
                data=generate_txt(df.drop(columns=["whatsapp_url"], errors="ignore")),
                file_name=f"{slug}.txt",
                mime="text/plain",
                width='stretch',
            )

        # ── Score explanation ──────────────────────────────────────────────────
        with st.expander("ℹ️ How is the Score calculated?"):
            st.markdown(
                """
                **Fitness formula for web dev / digital marketing services:**
                ```
                score_base  = (rating × 10) × (number_of_reviews / 100)
                score_final = score_base × 10   →  no website at all  (🔥 highest priority)
                score_final = score_base × 5    →  social network URL  (🟡 medium priority)
                score_final = score_base × 1    →  proper website       (low priority)
                ```
                - **No website** → green rows in Excel. Established businesses with zero digital presence.
                - **Social network only** → yellow rows. They have some online presence but no dedicated site.
                - **Has a page** → no highlight. Lowest priority for web dev / marketing services.

                > **Tip:** social-network-only profiles (Instagram, Facebook, LinkedIn) are **not**
                > counted as a real website — they score between "no site" and "has page".
                """
            )

    else:
        # ── Welcome / home screen ────────────────────────────────────────────────
        st.markdown(
            """
            ### How to use PyProspector

            | Step | Action |
            |:-----:|------|
            | 1️⃣ | Fill in the **niche** in the sidebar (e.g., *law firms*) |
            | 2️⃣ | Enter the **city/country**  (e.g., *Belo Horizonte, Brazil*) |
            | 3️⃣ | Adjust the maximum number of results and the minimum rating |
            | 4️⃣ | Click **🚀 Prospect Leads** and wait |

            ---

            **Ideal use cases:**
            - Web design agencies prospecting SMBs without a digital presence
            - SEO and paid traffic freelancers looking for new clients
            - Digital marketing consultants doing outbound B2B prospecting
            - Sales teams conducting local market research

            ---

            > ⚠️ **Responsible use:** use this tool in moderation and
            > respect the [Google Maps Terms of Service](https://maps.google.com/help/terms_maps/).
            > This project is for educational and ethical automation purposes.
            """
        )


if __name__ == "__main__":
    main()
